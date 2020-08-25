import openpyxl
import requests
# def writr_result(filename,sheetname,row,column,final_result):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb[sheetname]
#     sheet.cell(row=row, column=column).value = final_result
#     wb.save(filename)

def funr(url,res_body):
    request_header = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    res = requests.post(url = url,json = res_body,headers = request_header)
   # print(res.json())
    res_log = res.json()
    return res_log

def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    max_row = sheet.max_row #获取sheet里最大的行数
#print(max_row)
    case_list = []
    for i in range(2,max_row + 1):
        dict1 = dict(
        id = sheet.cell(row=i,column=1).value, #取的是测试用例编号
        url = sheet.cell(row=i,column=5).value,  #取得是url
        data = sheet.cell(row=i,column=6).value, #取得是data
        expect = sheet.cell(row=i,column=7).value ) #取得是预期结果
        case_list.append(dict1) #用append 把字典追加到列表里去 --》 列表就存放了所有的测试数据
    #print(case_list[0])
    #print(dict1)
    #print(case_list)
    return case_list #设置返回值，给别人去用
def writr_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value = final_result
    wb.save(filename)
def exeture_funx(filename,sheetname):
    res = read_data(filename,sheetname)
    #print(res)
    for testcase in res:      # 取出一条一条的测试用例
        case_id = testcase.get('id')   #字典取值或者value 取出id
        url = testcase.get('url')
        data = testcase.get('data')
        data = eval(data)    #运行被字符串包括的python表达式
        expect = testcase.get('expect')
        expect = eval(expect)  #把字符串转换成字典
        expect_msg = expect.get('msg')  #从预期结构中把msg取出来
        #print(case_id,url,expect)
        res1 = funr(url =url,res_body = data)  #调用发送请求的函数，并传入参数
        #print(res1)
        real_msg = res1.get('msg')
        print('预期结果为:{}'.format(expect_msg))
        print('实际结果为:{}'.format(real_msg))
        if real_msg == expect_msg:
            print('这条测试用例执行通过')
            final_res = '通过'
        else:
            print('这条测试用例执行不通过')
            final_res = '不通过，有bug'
        print('*'*100)
        writr_result(filename,sheetname,case_id+1,8,final_res)
exeture_funx('test_case_api.xlsx','register')
exeture_funx('test_case_api.xlsx','login')